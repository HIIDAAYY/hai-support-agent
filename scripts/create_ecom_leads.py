import openpyxl
import pandas as pd
import os

# Paths
excel_path = "C:/Users/aditm/OneDrive/Documents/Claude/Projects/5 million first/HAI-Outreach-Tracker.xlsx"
csv_path = "C:/Users/aditm/OneDrive/Documents/Claude/Projects/5 million first/Track-D-Ecom-Prospek.csv"

# Real Indonesian modest fashion & hijab brand owners research data
ecom_leads = [
  {
    "No": 1,
    "Brand Name": "Buttonscarves",
    "Owner/Founder": "Linda Anggrea",
    "Owner IG (Personal)": "@lindamario",
    "Brand IG": "@buttonscarves",
    "Website/Ecom": "buttonscarves.com",
    "Notes": "Modest fashion raksasa Indonesia. Owner-driven, engagement sangat tinggi, ekspansi global. Sangat cocok untuk otomatisasi chat support premium."
  },
  {
    "No": 2,
    "Brand Name": "Vanilla Hijab",
    "Owner/Founder": "Atina Maulia (Founder) & Intan Kusuma Fauzia (Co-Founder & CEO)",
    "Owner IG (Personal)": "@atinamaulia / @intankusumafauzia",
    "Brand IG": "@vanillahijab",
    "Website/Ecom": "vanillahijab.com",
    "Notes": "Brand hijab e-commerce native dengan traffic flash sale masif di website/WA. AI chatbot penanganan order & stok akan menghemat biaya CS operasional secara signifikan."
  },
  {
    "No": 3,
    "Brand Name": "Heaven Lights",
    "Owner/Founder": "Jihan Malik & Emma Malik",
    "Owner IG (Personal)": "@jihanmalik",
    "Brand IG": "@heaven_lights",
    "Website/Ecom": "heavenlights.co",
    "Notes": "Spesialis flash sale cepat sold out. Butuh chatbot instan saat puncak promo untuk memulihkan pesanan yang batal/pending (lost cart recovery)."
  },
  {
    "No": 4,
    "Brand Name": "Lozy Hijab",
    "Owner/Founder": "Andani Agni Putri",
    "Owner IG (Personal)": "@andaniagni",
    "Brand IG": "@lozyhijab",
    "Website/Ecom": "lozyhijab.com",
    "Notes": "Menargetkan pasar anak muda/Gen Z. Memiliki chat volume sangat tinggi di DM Instagram & WhatsApp. Sangat butuh response time instan 24/7."
  },
  {
    "No": 5,
    "Brand Name": "Wearing Klamby",
    "Owner/Founder": "Nadine Gaus",
    "Owner IG (Personal)": "@nadinegaus",
    "Brand IG": "@wearingklamby",
    "Website/Ecom": "klamby.id",
    "Notes": "Brand modest wear premium dengan storytelling budaya lokal sangat kuat. Integrasi chatbot mempermudah customer experience untuk cek ongkir & katalog produk."
  },
  {
    "No": 6,
    "Brand Name": "Kami. (Kami Idea)",
    "Owner/Founder": "Istafiana Candarini (Irin)",
    "Owner IG (Personal)": "@irin_kamidea",
    "Brand IG": "@kamiidea",
    "Website/Ecom": "kamidea.com",
    "Notes": "Brand pioneer hijab print & motif unik. Membutuhkan AI support untuk handling tanya jawab reseller dan program kemitraan."
  },
  {
    "No": 7,
    "Brand Name": "Meccanism",
    "Owner/Founder": "Zaskia Adya Mecca & Tasya Nur Medina",
    "Owner IG (Personal)": "@zaskiadyamecca / @tasyanurmedina",
    "Brand IG": "@meccanismbyzaskiamecca",
    "Website/Ecom": "meccanism.co.id",
    "Notes": "Artis-owned brand dengan ribuan followers aktif. Chatbot memproses info ketersediaan stok & panduan dropship."
  },
  {
    "No": 8,
    "Brand Name": "Elzatta Hijab",
    "Owner/Founder": "Elidawati Ali Oemar",
    "Owner IG (Personal)": "@elidawati.alioemar",
    "Brand IG": "@elzattahijab",
    "Website/Ecom": "elzatta.com",
    "Notes": "Salah satu pionir hijab instan terbesar di Indonesia dengan puluhan cabang offline. AI chatbot sangat pas untuk menyinkronkan pertanyaan ketersediaan stok cabang."
  }
]

# Convert to DataFrame
df = pd.DataFrame(ecom_leads)

# 1. Save to Excel as a new sheet
if os.path.exists(excel_path):
  try:
    print(f"Loading Excel workbook from {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    
    # Remove existing sheet if it exists to avoid duplicate
    sheet_name = "Track D - Ecom"
    if sheet_name in wb.sheetnames:
      print(f"   Removing existing sheet '{sheet_name}'...")
      wb.remove(wb[sheet_name])
      
    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)
    
    # Write DataFrame to sheet
    # Write headers
    headers = list(df.columns)
    ws.append(headers)
    
    # Write rows
    for r in df.itertuples(index=False):
      ws.append(list(r))
      
    # Save workbook
    wb.save(excel_path)
    print(f"Excel sheet '{sheet_name}' successfully created/updated!")
  except Exception as e:
    print(f"Error updating Excel: {e}")
else:
  print(f"Excel file not found at {excel_path}, skipped sheet creation.")

# 2. Save to CSV
try:
  df.to_csv(csv_path, index=False)
  print(f"CSV file successfully created at {csv_path}!")
except Exception as e:
  print(f"Error writing CSV: {e}")
